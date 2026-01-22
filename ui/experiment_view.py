import customtkinter as ctk
from pathlib import Path
from tkinter import filedialog
from openpyxl import load_workbook
import csv
import json
import traceback
try:
    import yaml
except ImportError:
    yaml = None
# pandas optional: not required for current readers (openpyxl/csv used)
pd = None


def _format_error(e: Exception) -> str:
    """Format exception for display in status label."""
    return f"❌ {e.__class__.__name__}: {e}"


def _log_error(e: Exception, context: str = ""):
    """Log error with traceback to stderr."""
    import sys
    if context:
        print(f"[ERROR] {context}", file=sys.stderr)
    traceback.print_exception(type(e), e, e.__traceback__, file=sys.stderr)


class ExperimentSection(ctk.CTkFrame):
    def __init__(self, master, on_change=None, on_send=None, on_minio_toggle=None):
        super().__init__(master, corner_radius=12)
        self.on_change = on_change
        self.on_send = on_send
        self.on_minio_toggle = on_minio_toggle
        self._selected_files: dict[str, set[str]] = {}
        self._metrics_settings: dict = {
            "header": True,
            "has_time": False,
            "time_col": "",
            "selected_cols": set(),
        }
        self._config_settings: dict = {
            "flatten": False,
            "use_custom_path": False,
            "custom_path": "",
            "parse_from_folder": False,
            "folder_pattern": "",
        }
        self._raw_data_settings: dict = {
            "send_minio": True,
            "save_locally": False,
            "local_path": "",
        }
        # CSV separators per selector (persisted)
        self._csv_separators: dict[str, str] = {
            "config": ",",
            "metrics": ",",
            "results": ",",
        }
        # batch sending controls (not persisted)
        self._batch_enable = False
        self._batch_selected: set[str] = set()
        self._allowed_tabular_suffixes = (".json", ".csv", ".xlsx", ".xlsm", ".yaml", ".yml")
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=0)
        self.grid_rowconfigure(999, weight=1)

        ctk.CTkLabel(self, text="Experiment Files", font=("Segoe UI", 18, "bold")).grid(
            row=0, column=0, columnspan=3, sticky="w", padx=12, pady=(12, 8)
        )

        ctk.CTkLabel(self, text="Experiment folder").grid(row=1, column=0, sticky="w", padx=12)
        self.folder_entry = ctk.CTkEntry(self, placeholder_text="Select a folder…")
        self.folder_entry.grid(row=1, column=1, sticky="ew", padx=6, pady=6)
        ctk.CTkButton(self, text="Browse…", command=self.choose_folder).grid(
            row=1, column=2, sticky="e", padx=(6, 12), pady=6
        )

        # --- Config section in a card (row 2) ---
        self.config_card = ctk.CTkFrame(self, corner_radius=10)
        self.config_card.grid(row=2, column=0, columnspan=3, sticky="ew", padx=12, pady=6)
        self.config_card.grid_columnconfigure(1, weight=1)
        
        config_title_frame = ctk.CTkFrame(self.config_card, fg_color="transparent")
        config_title_frame.grid(row=0, column=0, columnspan=3, sticky="w", padx=10, pady=(8, 6))
        ctk.CTkLabel(config_title_frame, text="Config", font=("Segoe UI", 14, "bold")).pack(side="left")
        ctk.CTkButton(
            config_title_frame, text="ℹ", width=20, height=20,
            fg_color="gray", hover_color="#5a5a5a",
            font=("Segoe UI", 11),
            command=lambda: self._show_info_tooltip("config")
        ).pack(side="left", padx=(6, 0))
        
        # Radio buttons for source selection
        self.config_source_var = ctk.StringVar(value="folder")
        
        def on_config_source_change():
            self._config_settings["use_custom_path"] = (self.config_source_var.get() == "custom")
            self._update_config_selector_visibility()
            self.render_details_sections()
            if callable(self.on_change):
                self.on_change()
        
        source_frame = ctk.CTkFrame(self.config_card, fg_color="transparent")
        source_frame.grid(row=1, column=0, columnspan=3, sticky="w", padx=10, pady=(0, 4))
        
        ctk.CTkRadioButton(
            source_frame, text="From folder", variable=self.config_source_var, 
            value="folder", command=on_config_source_change
        ).pack(side="left", padx=(0, 16))
        
        ctk.CTkRadioButton(
            source_frame, text="Custom path", variable=self.config_source_var,
            value="custom", command=on_config_source_change
        ).pack(side="left")
        
        # Config file dropdown (for "From folder" mode)
        self.config_file_menu = ctk.CTkOptionMenu(
            self.config_card, values=["None"], dynamic_resizing=False, width=300,
            command=lambda v: self.on_file_changed("config", v)
        )
        self.config_file_menu.grid(row=2, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 4))
        self.config_file_menu.set("None")
        
        # Config sheet menu (for Excel files)
        self.config_sheet_menu = ctk.CTkOptionMenu(
            self.config_card, values=[""], dynamic_resizing=False, width=150,
            command=lambda v: self.on_sheet_changed("config", v)
        )
        self.config_sheet_menu.grid(row=2, column=2, sticky="ew", padx=(6, 10), pady=(0, 4))
        self.config_sheet_menu.grid_remove()
        
        # Custom path entry + browse (for "Custom path" mode)
        self.config_custom_frame = ctk.CTkFrame(self.config_card, fg_color="transparent")
        self.config_custom_frame.grid(row=2, column=0, columnspan=3, sticky="ew", padx=10, pady=(0, 4))
        self.config_custom_frame.grid_columnconfigure(0, weight=1)
        self.config_custom_frame.grid_remove()  # Hidden by default
        
        self.config_custom_entry = ctk.CTkEntry(self.config_custom_frame, placeholder_text="Select a config file…")
        self.config_custom_entry.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        
        def choose_custom_config():
            filetypes = [
                ("Config files", "*.json *.yaml *.yml *.csv *.xlsx *.xlsm"),
                ("JSON", "*.json"),
                ("YAML", "*.yaml *.yml"),
                ("CSV", "*.csv"),
                ("Excel", "*.xlsx *.xlsm"),
                ("All files", "*.*"),
            ]
            filepath = filedialog.askopenfilename(filetypes=filetypes)
            if filepath:
                self.config_custom_entry.delete(0, "end")
                self.config_custom_entry.insert(0, filepath)
                self._config_settings["custom_path"] = filepath
                self.render_details_sections()
                if callable(self.on_change):
                    self.on_change()
        
        ctk.CTkButton(self.config_custom_frame, text="Browse…", width=90, command=choose_custom_config).grid(
            row=0, column=1, sticky="e"
        )
        
        # Parse config from folder name option
        self.parse_folder_frame = ctk.CTkFrame(self.config_card, fg_color="transparent")
        self.parse_folder_frame.grid(row=3, column=0, columnspan=3, sticky="w", padx=10, pady=(4, 8))
        
        self.parse_folder_var = ctk.BooleanVar(value=False)
        
        def on_parse_folder_toggle():
            self._config_settings["parse_from_folder"] = bool(self.parse_folder_var.get())
            self._update_parse_button_visibility()
            self.render_details_sections()
            if callable(self.on_change):
                self.on_change()
        
        ctk.CTkCheckBox(
            self.parse_folder_frame, text="Parse config from folder name", 
            variable=self.parse_folder_var, command=on_parse_folder_toggle
        ).pack(side="left")
        
        self.parse_pattern_btn = ctk.CTkButton(
            self.parse_folder_frame, text="Define pattern…", width=110, 
            command=self._open_folder_pattern_dialog
        )
        self.parse_pattern_btn.pack(side="left", padx=(8, 0))
        self.parse_pattern_btn.pack_forget()  # Hidden by default

        # Store references for the file_menus and sheet_menus dicts
        self.file_menus: dict[str, ctk.CTkOptionMenu] = {"config": self.config_file_menu}
        self.sheet_menus: dict[str, ctk.CTkOptionMenu] = {"config": self.config_sheet_menu}
        
        # --- Other selectors (results, metrics, raw_data, artifacts) ---
        self._keys = [
            ("config", "Config"),  # Keep for compatibility but handled separately above
            ("results", "Results"),
            ("metrics", "Metrics"),
            ("raw_data", "Raw data"),
            ("artifacts", "Artifacts"),
        ]
        
        # Info tooltips for selectors
        self._info_tooltips = {
            "config": (
                "📋 Experiment Configuration & Metadata\n\n"
                "Stores experiment parameters and settings.\n\n"
                "Accepted formats:\n"
                "  • JSON (.json) - nested or flat structure\n"
                "  • YAML (.yaml, .yml) - nested or flat\n"
                "  • CSV (.csv) - key-value pairs\n"
                "  • Excel (.xlsx, .xlsm) - key-value pairs\n\n"
                "Options:\n"
                "  • Flatten: converts nested keys to flat (a.b → a_b)\n"
                "  • Parse from folder: extract values from folder name"
            ),
            "metrics": (
                "📈 Metrics & Time Series Data\n\n"
                "Stores series of values logged during experiment.\n"
                "Values are logged as scalars in Sacred.\n\n"
                "Accepted formats:\n"
                "  • CSV (.csv) - columns = metrics, rows = values\n"
                "  • Excel (.xlsx, .xlsm) - same structure\n\n"
                "Options:\n"
                "  • Column header: first row contains column names\n"
                "  • X-axis column: use a column as step/time axis"
            ),
            "results": (
                "🏆 Experimental Results\n\n"
                "Stores final experiment outcomes and scores.\n"
                "These values appear in Sacred's result field.\n\n"
                "Accepted formats:\n"
                "  • JSON (.json) - key-value pairs\n"
                "  • CSV (.csv) - two columns (key, value)\n"
                "  • Excel (.xlsx, .xlsm) - two columns (key, value)"
            ),
            "raw_data": (
                "💾 Raw Data Files\n\n"
                "Large files sent to MinIO or saved locally.\n"
                "Use for files > 24 MB (MongoDB limit).\n\n"
                "Examples: datasets, model weights, logs..."
            ),
            "artifacts": (
                "📎 Artifacts\n\n"
                "Small files stored directly in MongoDB.\n"
                "Use for files < 24 MB.\n\n"
                "Examples: images, plots, small CSVs..."
            ),
        }
        
        # Start from row 3 for the other selectors (config is at row 2)
        for idx, (key, label) in enumerate(self._keys[1:], start=3):  # Skip config
            # Create label frame with info button for all selectors that have tooltips
            label_frame = ctk.CTkFrame(self, fg_color="transparent")
            label_frame.grid(row=idx, column=0, sticky="w", padx=12)
            ctk.CTkLabel(label_frame, text=label).pack(side="left")
            if key in self._info_tooltips:
                info_btn = ctk.CTkButton(
                    label_frame, text="ℹ", width=20, height=20, 
                    fg_color="gray", hover_color="#5a5a5a",
                    font=("Segoe UI", 11),
                    command=lambda k=key: self._show_info_tooltip(k)
                )
                info_btn.pack(side="left", padx=(4, 0))
            
            initial_values = ["None"]
            file_menu = ctk.CTkOptionMenu(
                self, values=initial_values, dynamic_resizing=False, width=320,
                command=lambda v, k=key: self.on_file_changed(k, v)
            )
            file_menu.grid(row=idx, column=1, sticky="ew", padx=6, pady=6)
            file_menu.set("None")
            self.file_menus[key] = file_menu

            sheet_menu = ctk.CTkOptionMenu(self, values=[""], dynamic_resizing=False, width=200, command=lambda v, k=key: self.on_sheet_changed(k, v))
            sheet_menu.grid(row=idx, column=2, sticky="ew", padx=(6, 12), pady=6)
            sheet_menu.grid_remove()
            self.sheet_menus[key] = sheet_menu

        # container for dynamic per-selector sections (two-column layout, small margins)
        self.details_container = ctk.CTkFrame(self, corner_radius=12)
        # Row 0: title, Row 1: folder, Row 2: config, Rows 3-6: other selectors (4 items)
        base_row = 3 + len(self._keys[1:]) + 1  # = 3 + 4 + 1 = 8
        self.details_container.grid(row=base_row, column=0, columnspan=3, sticky="nsew", padx=6, pady=(6, 6))
        self.details_container.grid_columnconfigure(0, weight=1)
        self.details_container.grid_columnconfigure(1, weight=1)

        # Batch toggle and siblings list
        batch_row = base_row + 1
        self.batch_enable_var = ctk.BooleanVar(value=False)
        def on_batch_toggle():
            self._batch_enable = bool(self.batch_enable_var.get())
            self._render_batch_checkboxes()
            if callable(self.on_change):
                self.on_change()
        ctk.CTkCheckBox(self, text="Send multiple experiments", variable=self.batch_enable_var, command=on_batch_toggle).grid(
            row=batch_row, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 4)
        )
        self.batch_container = ctk.CTkFrame(self, corner_radius=8)
        self.batch_container.grid(row=batch_row + 1, column=0, columnspan=3, sticky="nsew", padx=6, pady=(0, 6))
        # hide container by default so it doesn't reserve space
        try:
            self.batch_container.grid_remove()
        except Exception:
            pass
        self.batch_container.grid_columnconfigure(0, weight=1)
        self._render_batch_checkboxes()

        # actions row: Send experiment button inside the section, below batch
        actions_row = ctk.CTkFrame(self, fg_color="transparent")
        actions_row.grid(row=batch_row + 2, column=0, columnspan=3, sticky="ew", padx=6, pady=(0, 2))
        actions_row.grid_columnconfigure(2, weight=1)
        self.send_btn = ctk.CTkButton(actions_row, text="Send experiment", width=180, height=36, command=self._on_send_click)
        self.send_btn.grid(row=0, column=2, sticky="e", padx=(0, 6), pady=(2, 2))

        # status labels: one for file/cards errors, one for send result
        self.status = ctk.CTkLabel(self, text="", wraplength=520, justify="left")
        self.status.grid(row=batch_row + 3, column=0, columnspan=3, sticky="ew", padx=12, pady=(2, 6))
        self.send_status = ctk.CTkLabel(self, text="", wraplength=520, justify="left")
        self.send_status.grid(row=batch_row + 4, column=0, columnspan=3, sticky="ew", padx=12, pady=(2, 4))

    def _on_send_click(self):
        try:
            if callable(self.on_send):
                self.on_send()
        except Exception as e:
            _log_error(e, "Error in send click handler")
            self.send_status.configure(text=_format_error(e))

    # --- IO ---
    def get_prefs(self) -> dict:
        data = {"experiment_folder": self.folder_entry.get().strip()}
        for key, _ in self._keys:
            data[f"{key}_name"] = (self.file_menus[key].get() or "").strip()
            data[f"{key}_sheet"] = (self.sheet_menus[key].get() or "").strip()
            if key in ("raw_data", "artifacts"):
                selected = sorted(list(self._selected_files.get(key, set())))
                data[f"{key}_files"] = selected
        # metrics settings persistence
        data["metrics_header"] = int(bool(self._metrics_settings.get("header", True)))
        data["metrics_has_time"] = int(bool(self._metrics_settings.get("has_time", False)))
        data["metrics_time_col"] = self._metrics_settings.get("time_col", "")
        data["metrics_selected_cols"] = sorted(list(self._metrics_settings.get("selected_cols", set())))
        # config settings persistence
        data["config_flatten"] = int(bool(self._config_settings.get("flatten", False)))
        data["config_use_custom_path"] = int(bool(self._config_settings.get("use_custom_path", False)))
        data["config_custom_path"] = self._config_settings.get("custom_path", "")
        data["config_parse_from_folder"] = int(bool(self._config_settings.get("parse_from_folder", False)))
        data["config_folder_pattern"] = self._config_settings.get("folder_pattern", "")
        # Include parsed folder values if enabled
        if self._config_settings.get("parse_from_folder", False):
            data["config_parsed_folder_values"] = self._parse_folder_name()
        else:
            data["config_parsed_folder_values"] = {}
        # raw_data settings persistence
        data["raw_data_send_minio"] = int(bool(self._raw_data_settings.get("send_minio", True)))
        data["raw_data_save_locally"] = int(bool(self._raw_data_settings.get("save_locally", False)))
        data["raw_data_local_path"] = self._raw_data_settings.get("local_path", "")
        # CSV separators
        data["config_sep"] = self._csv_separators.get("config", ",")
        data["metrics_sep"] = self._csv_separators.get("metrics", ",")
        data["results_sep"] = self._csv_separators.get("results", ",")
        # compute list of folders per batch toggle
        folders_list: list[str] = []
        base_folder = (self.folder_entry.get() or "").strip()
        try:
            if self._batch_enable:
                # list siblings of selected folder
                if base_folder:
                    p = Path(base_folder)
                    parent = p.parent if p.exists() else None
                    if parent and parent.exists():
                        # if no selection yet, default to all siblings
                        if not self._batch_selected:
                            for d in parent.iterdir():
                                if d.is_dir() and not d.name.startswith("."):
                                    self._batch_selected.add(d.name)
                        for name in sorted(list(self._batch_selected)):
                            folders_list.append(str((parent / name).resolve()))
            else:
                if base_folder:
                    folders_list = [str(Path(base_folder).resolve())]
        except Exception as e:
            _log_error(e, "Error computing experiment folders list")
            folders_list = [base_folder] if base_folder else []
        data["experiment_folders"] = folders_list
        return data

    def set_prefs(self, data: dict):
        exp_dir = data.get("experiment_folder", "")
        if exp_dir:
            self.folder_entry.delete(0, "end")
            self.folder_entry.insert(0, exp_dir)
        # populate menu values
        self.refresh_items(preserve_selection=False)
        # restore selection
        for key, _ in self._keys:
            name = data.get(f"{key}_name", "") or "None"
            try:
                self.file_menus[key].set(name)
            except Exception:
                self.file_menus[key].set(name)
            self.update_sheet_menu_for(key)
        # restore sheet
        for key, _ in self._keys:
            sheet = data.get(f"{key}_sheet", "")
            if sheet:
                try:
                    self.sheet_menus[key].set(sheet)
                except Exception:
                    self.sheet_menus[key].set(sheet)
            self.update_sheet_menu_for(key)
        # restore per-folder file selections
        for key in ("raw_data", "artifacts"):
            saved = data.get(f"{key}_files", [])
            if isinstance(saved, list):
                self._selected_files[key] = set(saved)
        # restore metrics settings
        self._metrics_settings["header"] = bool(data.get("metrics_header", 1))
        self._metrics_settings["has_time"] = bool(data.get("metrics_has_time", 0))
        self._metrics_settings["time_col"] = data.get("metrics_time_col", "") or ""
        sel = data.get("metrics_selected_cols", [])
        self._metrics_settings["selected_cols"] = set(sel) if isinstance(sel, list) else set()
        # restore config settings
        self._config_settings["flatten"] = bool(data.get("config_flatten", 0))
        self._config_settings["use_custom_path"] = bool(data.get("config_use_custom_path", 0))
        self._config_settings["custom_path"] = data.get("config_custom_path", "") or ""
        self._config_settings["parse_from_folder"] = bool(data.get("config_parse_from_folder", 0))
        self._config_settings["folder_pattern"] = data.get("config_folder_pattern", "") or ""
        # Update config source radio button and visibility
        self.config_source_var.set("custom" if self._config_settings["use_custom_path"] else "folder")
        self._update_config_selector_visibility()
        # Update parse from folder checkbox and button visibility
        self.parse_folder_var.set(self._config_settings["parse_from_folder"])
        self._update_parse_button_visibility()
        # restore raw_data settings
        self._raw_data_settings["send_minio"] = bool(data.get("raw_data_send_minio", 1))
        self._raw_data_settings["save_locally"] = bool(data.get("raw_data_save_locally", 0))
        self._raw_data_settings["local_path"] = data.get("raw_data_local_path", "") or ""
        # restore CSV separators
        self._csv_separators["config"] = data.get("config_sep", ",") or ","
        self._csv_separators["metrics"] = data.get("metrics_sep", ",") or ","
        self._csv_separators["results"] = data.get("results_sep", ",") or ","
        # restore experiment name
        # self.exp_name_entry.delete(0, "end")
        # self.exp_name_entry.insert(0, data.get("experiment_name", ""))
        # ensure details reflect restored prefs
        self.render_details_sections()

    def _render_batch_checkboxes(self):
        # clear
        for child in list(self.batch_container.winfo_children() if hasattr(self, 'batch_container') else []):
            child.destroy()
        if not getattr(self, 'batch_container', None):
            return
        # Hide the container when disabled; show when enabled
        if not bool(self.batch_enable_var.get()):
            self.batch_container.grid_remove()
            return
        else:
            self.batch_container.grid()
        # build siblings list
        base_folder = (self.folder_entry.get() or "").strip()
        siblings: list[str] = []
        try:
            if base_folder:
                p = Path(base_folder)
                parent = p.parent if p.exists() else None
                if parent and parent.exists():
                    siblings = [d.name for d in parent.iterdir() if d.is_dir() and not d.name.startswith('.')]
                    siblings.sort(key=lambda n: n.lower())
        except Exception as e:
            _log_error(e, "Error listing sibling folders")
            siblings = []
        # default select all if nothing yet
        if not self._batch_selected:
            self._batch_selected = set(siblings)
        # render
        for i, name in enumerate(siblings):
            var = ctk.BooleanVar(value=(name in self._batch_selected))
            def _toggle(n=name, v=var):
                if v.get():
                    self._batch_selected.add(n)
                else:
                    self._batch_selected.remove(n)
                if callable(self.on_change):
                    self.on_change()
            cb = ctk.CTkCheckBox(self.batch_container, text=name, variable=var, command=_toggle)
            cb.grid(row=i, column=0, sticky="w", padx=8, pady=2)

    # --- Folder pattern parsing ---
    def _parse_folder_name(self) -> dict:
        """Parse the folder name using the defined pattern and return extracted values.
        
        Pattern syntax:
        - $variable$ - matches any characters (greedy minimal)
        - $variable%N$ - matches exactly N characters (e.g., $date%8$ matches 8 chars)
        """
        pattern = self._config_settings.get("folder_pattern", "")
        if not pattern:
            return {}
        
        folder_path = self.folder_entry.get().strip()
        if not folder_path:
            return {}
        
        folder_name = Path(folder_path).name
        
        # Extract variable definitions from pattern (between $ signs)
        # Format: $name$ or $name%length$
        import re
        var_pattern = r'\$([^$%]+)(?:%(\d+))?\$'
        variables = re.findall(var_pattern, pattern)
        
        if not variables:
            return {}
        
        # Build regex from pattern by replacing $var$ or $var%N$ with named groups
        regex_pattern = re.escape(pattern)
        for var_name, length in variables:
            if length:
                # Exact length: match exactly N characters
                escaped_var = re.escape(f"${var_name}%{length}$")
                regex_pattern = regex_pattern.replace(escaped_var, f"(?P<{var_name}>.{{{length}}})")
            else:
                # No length: match any characters (non-greedy)
                escaped_var = re.escape(f"${var_name}$")
                regex_pattern = regex_pattern.replace(escaped_var, f"(?P<{var_name}>.+?)")
        
        # Try to match
        try:
            match = re.match(f"^{regex_pattern}$", folder_name)
            if match:
                return match.groupdict()
        except re.error:
            pass
        
        return {}

    def _open_folder_pattern_dialog(self):
        """Open a dialog to define the folder name pattern."""
        folder_path = self.folder_entry.get().strip()
        folder_name = Path(folder_path).name if folder_path else "(no folder selected)"
        
        dialog = ctk.CTkToplevel(self)
        dialog.title("Define Folder Pattern")
        dialog.geometry("550x520")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center on parent
        dialog.update_idletasks()
        x = self.winfo_rootx() + 50
        y = self.winfo_rooty() + 50
        dialog.geometry(f"+{x}+{y}")
        
        dialog.grid_columnconfigure(0, weight=1)
        dialog.grid_rowconfigure(6, weight=1)
        
        # Explanation section
        help_frame = ctk.CTkFrame(dialog, corner_radius=8)
        help_frame.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 12))
        help_frame.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(help_frame, text="📝 Pattern Syntax", font=("Segoe UI", 12, "bold")).grid(
            row=0, column=0, sticky="w", padx=10, pady=(8, 4)
        )
        
        help_text = (
            "Use variables between $ signs to extract values from folder name:\n\n"
            "  $variable$      → matches any text (flexible length)\n"
            "  $variable%N$   → matches exactly N characters\n\n"
            "Examples:\n"
            "  Folder:    MyExp_20260121_v2\n"
            "  Pattern:   $name$_$date%8$_v$ver$\n"
            "  Result:     name=MyExp, date=20260121, ver=2"
        )
        help_label = ctk.CTkLabel(help_frame, text=help_text, font=("Consolas", 11), justify="left")
        help_label.grid(row=1, column=0, sticky="w", padx=10, pady=(0, 8))
        
        # Folder name display
        ctk.CTkLabel(dialog, text="Current folder name:", font=("Segoe UI", 12, "bold")).grid(
            row=1, column=0, sticky="w", padx=16, pady=(8, 4)
        )
        folder_label = ctk.CTkLabel(dialog, text=folder_name, font=("Consolas", 12))
        folder_label.grid(row=2, column=0, sticky="w", padx=16, pady=(0, 12))
        
        # Pattern input
        ctk.CTkLabel(dialog, text="Your pattern:", font=("Segoe UI", 12, "bold")).grid(
            row=3, column=0, sticky="w", padx=16, pady=(8, 4)
        )
        pattern_entry = ctk.CTkEntry(dialog, placeholder_text="e.g., $name$_$date%8$_v$version$")
        pattern_entry.grid(row=4, column=0, sticky="ew", padx=16, pady=(0, 12))
        
        # Restore current pattern
        current_pattern = self._config_settings.get("folder_pattern", "")
        if current_pattern:
            pattern_entry.insert(0, current_pattern)
        
        # Extracted values preview
        ctk.CTkLabel(dialog, text="Extracted values:", font=("Segoe UI", 12, "bold")).grid(
            row=5, column=0, sticky="nw", padx=16, pady=(8, 4)
        )
        preview_box = ctk.CTkTextbox(dialog, height=100, wrap="none", font=("Consolas", 11))
        preview_box.grid(row=6, column=0, sticky="nsew", padx=16, pady=(0, 12))
        
        def update_preview(*args):
            pattern = pattern_entry.get()
            self._config_settings["folder_pattern"] = pattern
            parsed = self._parse_folder_name()
            
            preview_box.configure(state="normal")
            preview_box.delete("1.0", "end")
            
            if parsed:
                preview_text = "\n".join(f"{k}: {v}" for k, v in parsed.items())
                preview_box.insert("1.0", preview_text)
            else:
                if pattern:
                    preview_box.insert("1.0", "(no match - check your pattern)")
                else:
                    preview_box.insert("1.0", "(enter a pattern above)")
            
            preview_box.configure(state="disabled")
        
        # Update preview on typing
        pattern_entry.bind("<KeyRelease>", update_preview)
        update_preview()  # Initial update
        
        # Buttons
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.grid(row=7, column=0, sticky="e", padx=16, pady=(0, 16))
        
        def save_and_close():
            self._config_settings["folder_pattern"] = pattern_entry.get()
            self.render_details_sections()
            if callable(self.on_change):
                self.on_change()
            dialog.destroy()
        
        ctk.CTkButton(btn_frame, text="Cancel", width=80, fg_color="gray", command=dialog.destroy).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_frame, text="Save", width=80, command=save_and_close).pack(side="left")
        
        dialog.bind("<Escape>", lambda e: dialog.destroy())
        pattern_entry.focus_set()

    # --- Info tooltip display ---
    def _show_info_tooltip(self, key: str):
        """Show an info dialog for the given selector key."""
        info_text = self._info_tooltips.get(key, "")
        if not info_text:
            return
        
        # Create popup window - size based on content
        popup = ctk.CTkToplevel(self)
        popup.title("Info")
        popup.transient(self)
        popup.grab_set()
        
        popup.grid_columnconfigure(0, weight=1)
        popup.grid_rowconfigure(0, weight=1)
        
        # Message in a frame with padding
        msg_frame = ctk.CTkFrame(popup, corner_radius=8)
        msg_frame.grid(row=0, column=0, padx=16, pady=(16, 8), sticky="nsew")
        
        msg_label = ctk.CTkLabel(msg_frame, text=info_text, wraplength=350, justify="left", font=("Consolas", 11))
        msg_label.pack(padx=12, pady=12)
        
        # OK button
        ok_btn = ctk.CTkButton(popup, text="OK", width=80, command=popup.destroy)
        ok_btn.grid(row=1, column=0, pady=(0, 12))
        
        # Update to calculate size
        popup.update_idletasks()
        
        # Set geometry based on content (min 400x200)
        width = max(400, popup.winfo_reqwidth())
        height = max(180, popup.winfo_reqheight())
        popup.geometry(f"{width}x{height}")
        
        # Center on parent
        x = self.winfo_rootx() + 100
        y = self.winfo_rooty() + 50
        popup.geometry(f"+{x}+{y}")
        
        # Close on escape
        popup.bind("<Escape>", lambda e: popup.destroy())
        popup.focus_set()

    # --- Config source visibility ---
    def _update_config_selector_visibility(self):
        """Show either the folder dropdown or custom path entry based on source selection."""
        use_custom = self._config_settings.get("use_custom_path", False)
        if use_custom:
            self.config_file_menu.grid_remove()
            self.config_sheet_menu.grid_remove()
            self.config_custom_frame.grid()
            # Restore custom path if previously set
            if self._config_settings.get("custom_path"):
                self.config_custom_entry.delete(0, "end")
                self.config_custom_entry.insert(0, self._config_settings.get("custom_path", ""))
        else:
            self.config_custom_frame.grid_remove()
            self.config_file_menu.grid()
            # Sheet menu visibility handled by update_sheet_menu_for

    def _update_parse_button_visibility(self):
        """Show or hide the parse pattern button based on checkbox state."""
        if self._config_settings.get("parse_from_folder", False):
            self.parse_pattern_btn.pack(side="left", padx=(8, 0))
        else:
            self.parse_pattern_btn.pack_forget()

    # --- Events ---
    def choose_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.folder_entry.delete(0, "end")
            self.folder_entry.insert(0, folder)
            self.refresh_items(preserve_selection=True)
        if callable(self.on_change):
            self.on_change()

    def on_file_changed(self, key: str, value: str):
        # Reset dependent selections when source file/folder changes
        if key == "metrics":
            # clear selected columns and time column; will be recomputed on render
            self._metrics_settings["selected_cols"] = set()
            self._metrics_settings["time_col"] = ""
        elif key in ("raw_data", "artifacts"):
            # clear previously selected files so defaults (all files) apply
            self._selected_files[key] = set()

        self.update_sheet_menu_for(key)
        self.render_details_sections()
        if callable(self.on_change):
            self.on_change()

    def on_sheet_changed(self, key: str, value: str):
        # When a sheet is selected, simply re-render the cards
        self.render_details_sections()
        if callable(self.on_change):
            self.on_change()

    def get_full_path_for_key(self, key: str) -> Path | None:
        # Special handling for config with custom path
        if key == "config" and self._config_settings.get("use_custom_path", False):
            custom_path = self._config_settings.get("custom_path", "")
            if custom_path:
                return Path(custom_path)
            return None
        
        base_folder = self.folder_entry.get().strip()
        name = self.file_menus[key].get().strip()
        if not base_folder or not name or name == "None":
            return None
        return Path(base_folder) / name

    def update_sheet_menu_for(self, key: str):
        sheet_menu = self.sheet_menus[key]
        path = self.get_full_path_for_key(key)
        # hide if not supported tabular file
        # special-case: for raw_data and artifacts never show sheet selector
        if not path or path.is_dir() or key in ("raw_data", "artifacts") or path.suffix.lower() not in (".xlsx", ".xlsm"):
            sheet_menu.grid_remove()
            sheet_menu.configure(values=[""])
            sheet_menu.set("")
            return
        # fetch sheet names
        sheets: list[str] = []
        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            sheets = list(wb.sheetnames)
            wb.close()
        except Exception as e:
            self.status.configure(text=f"Could not read sheets from {path.name}: {e}")
        if not sheets:
            sheets = [""]
        
        sheet_menu.configure(values=sheets)
        sheet_menu.grid()
        # show and keep previous selection if still present
        current = sheet_menu.get() or ""
        if current and current in sheets:
            sheet_menu.set(current)
        else:
            sheet_menu.set(sheets[0] if sheets else "")

    def refresh_items(self, preserve_selection: bool = True):
        base_folder = self.folder_entry.get().strip()
        all_items = self._list_items()
        restricted = {"config", "metrics", "results"}
        for key, _ in self._keys:
            current = (self.file_menus[key].get() or "") if preserve_selection else ""
            # Build values list per key
            if key in restricted and base_folder:
                try:
                    base = Path(base_folder)
                    filtered = [n for n in all_items if (base / n).is_file() and (base / n).suffix.lower() in self._allowed_tabular_suffixes]
                except Exception as e:
                    _log_error(e, f"Error filtering files for {key}")
                    self.status.configure(text=_format_error(e))
                    filtered = []
                values = ["None"] + filtered
            else:
                values = ["None"] + list(all_items)

            # Configure menu
            self.file_menus[key].configure(values=values)

            # Reset invalid current selections
            target_value = current if (current and current in values) else "None"
            self.file_menus[key].set(target_value)

        # refresh details after items update
        self.render_details_sections()

    def _list_items(self) -> list[str]:
        base_folder = self.folder_entry.get().strip()
        if not base_folder:
            return []
        base = Path(base_folder)
        if not base.exists() or not base.is_dir():
            return []
        names = [p.name for p in base.iterdir()]
        names.sort(key=lambda n: n.lower())
        return names

    def _render_config_card(self, idx: int, cols: int) -> int:
        """Render the config card with options and preview. Returns updated idx."""
        use_custom = self._config_settings.get("use_custom_path", False)
        parse_from_folder = self._config_settings.get("parse_from_folder", False)
        
        # Determine which path to use
        if use_custom:
            custom_path = self._config_settings.get("custom_path", "")
            path = Path(custom_path) if custom_path else None
            display_name = Path(custom_path).name if custom_path else "(none)"
        else:
            folder_name = (self.file_menus["config"].get() or "").strip()
            has_file = folder_name and folder_name != "None"
            if not has_file and not parse_from_folder:
                return idx  # No config selected and no folder parsing, don't show card
            path = self.get_full_path_for_key("config") if has_file else None
            display_name = folder_name if has_file else "(none)"
        
        has_valid_file = path and path.is_file()
        
        # Show card if we have a file OR if parse_from_folder is enabled
        if not has_valid_file and not parse_from_folder:
            return idx
        
        r, c = divmod(idx, cols)
        sec = ctk.CTkFrame(self.details_container, corner_radius=12)
        sec.grid(row=r, column=c, sticky="nsew", padx=6, pady=6)
        sec.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(sec, text="Config", font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=8, pady=(8, 4)
        )
        
        current_row = 1
        
        # Show selected file (only if we have one)
        if has_valid_file:
            ctk.CTkLabel(sec, text="File").grid(row=current_row, column=0, sticky="w", padx=8, pady=4)
            ctk.CTkLabel(sec, text=display_name).grid(row=current_row, column=1, sticky="w", padx=(6, 8), pady=4)
            current_row += 1
        
        # Sheet selector for Excel files
        sheet = ""
        if has_valid_file and path.suffix.lower() in (".xlsx", ".xlsm"):
            sheet = (self.sheet_menus["config"].get() or "").strip()
            if sheet:
                ctk.CTkLabel(sec, text="Sheet").grid(row=current_row, column=0, sticky="w", padx=8, pady=4)
                ctk.CTkLabel(sec, text=sheet).grid(row=current_row, column=1, sticky="w", padx=(6, 8), pady=4)
                current_row += 1
        
        # CSV separator
        if has_valid_file and path.suffix.lower() == ".csv":
            ctk.CTkLabel(sec, text="Separator").grid(row=current_row, column=0, sticky="w", padx=8, pady=4)
            sep_menu = ctk.CTkOptionMenu(
                sec, values=[",", ";", "|", "\\t"], dynamic_resizing=False,
                command=lambda v: self._on_sep_changed("config", v)
            )
            current_sep = self._csv_separators.get("config", ",")
            display_val = "\\t" if current_sep == "\t" else current_sep
            sep_menu.set(display_val)
            sep_menu.grid(row=current_row, column=1, sticky="ew", padx=(6, 8), pady=4)
            current_row += 1
        
        # Flatten checkbox for JSON/YAML
        if has_valid_file and path.suffix.lower() in (".json", ".yaml", ".yml"):
            flatten_var = ctk.BooleanVar(value=bool(self._config_settings.get("flatten", False)))
            def on_flatten_toggle():
                self._config_settings["flatten"] = bool(flatten_var.get())
                self.render_details_sections()
                if callable(self.on_change):
                    self.on_change()
            flatten_cb = ctk.CTkCheckBox(sec, text="Flatten nested keys", variable=flatten_var, command=on_flatten_toggle)
            flatten_cb.grid(row=current_row, column=0, columnspan=2, sticky="w", padx=8, pady=4)
            current_row += 1
        
        # Show parsed values if pattern is set (parse option is in main UI now)
        if self._config_settings.get("parse_from_folder", False) and self._config_settings.get("folder_pattern", ""):
            parsed = self._parse_folder_name()
            if parsed:
                ctk.CTkLabel(sec, text="Parsed values", font=("Segoe UI", 12, "bold")).grid(
                    row=current_row, column=0, columnspan=2, sticky="w", padx=8, pady=(8, 2)
                )
                current_row += 1
                
                parsed_text = "\n".join(f"{k}: {v}" for k, v in parsed.items())
                parsed_box = ctk.CTkTextbox(sec, height=60, width=300, wrap="none", font=("Consolas", 11))
                parsed_box.grid(row=current_row, column=0, columnspan=2, sticky="ew", padx=8, pady=(2, 4))
                parsed_box.insert("1.0", parsed_text)
                parsed_box.configure(state="disabled")
                current_row += 1
        
        # Preview (only if we have a valid file)
        if has_valid_file:
            ctk.CTkLabel(sec, text="Preview", font=("Segoe UI", 12, "bold")).grid(
                row=current_row, column=0, columnspan=2, sticky="w", padx=8, pady=(8, 2)
            )
            current_row += 1
            
            preview_text = self._read_config_preview(path, sheet)
            preview_box = ctk.CTkTextbox(sec, height=150, width=300, wrap="none", font=("Consolas", 11))
            preview_box.grid(row=current_row, column=0, columnspan=2, sticky="nsew", padx=8, pady=(2, 8))
            preview_box.insert("1.0", preview_text)
            preview_box.configure(state="disabled")
        
        return idx + 1

    # --- Dynamic details per selector ---
    def render_details_sections(self):
        # clear any previous error message on each cards update
        self.status.configure(text="")
        # clear previous
        for child in list(self.details_container.winfo_children()):
            child.destroy()
        cols = 2
        idx = 0
        
        # Always render config card first (it has special handling for custom path)
        idx = self._render_config_card(idx, cols)
        
        for key, label in self._keys:
            # Skip config - already rendered above
            if key == "config":
                continue
            name = (self.file_menus[key].get() or "").strip()
            if not name or name == "None":
                continue
            path = self.get_full_path_for_key(key)
            # create a simple frame per selection in a two-column grid
            r, c = divmod(idx, cols)
            sec = ctk.CTkFrame(self.details_container, corner_radius=12)
            sec.grid(row=r, column=c, sticky="nsew", padx=6, pady=6)
            sec.grid_columnconfigure(1, weight=1)
            ctk.CTkLabel(sec, text=f"{label}", font=("Segoe UI", 14, "bold")).grid(
                row=0, column=0, columnspan=2, sticky="w", padx=8, pady=(8, 4)
            )
            ctk.CTkLabel(sec, text="Selected file").grid(row=1, column=0, sticky="w", padx=8, pady=4)
            ctk.CTkLabel(sec, text=name).grid(row=1, column=1, sticky="w", padx=(6, 8), pady=4)
            # sheet (if visible)
            sheet = (self.sheet_menus[key].get() or "").strip()
            if sheet and not (key in ("raw_data", "artifacts")):
                ctk.CTkLabel(sec, text="Sheet").grid(row=2, column=0, sticky="w", padx=8, pady=4)
                ctk.CTkLabel(sec, text=sheet).grid(row=2, column=1, sticky="w", padx=(6, 8), pady=4)
            # CSV separator selector for config/metrics/results
            if path and path.is_file() and path.suffix.lower() == ".csv" and key in ("config", "metrics", "results"):
                sep_row = 3 if (sheet and not (key in ("raw_data", "artifacts"))) else 2
                ctk.CTkLabel(sec, text="Separator").grid(row=sep_row, column=0, sticky="w", padx=8, pady=4)
                sep_menu = ctk.CTkOptionMenu(
                    sec,
                    values=[",", ";", "|", "\\t"],
                    dynamic_resizing=False,
                    command=lambda v, k=key: self._on_sep_changed(k, v)
                )
                current_sep = self._csv_separators.get(key, ",")
                display_val = "\\t" if current_sep == "\t" else current_sep
                sep_menu.set(display_val)
                sep_menu.grid(row=sep_row, column=1, sticky="ew", padx=(6, 8), pady=4)
            # folder checklist for raw_data / artifacts
            if key in ("raw_data", "artifacts") and path and path.is_dir():
                files = []
                files = [p.name for p in Path(path).iterdir() if p.is_file()]
                files.sort(key=lambda n: n.lower())
                # initialize default selection: previously saved intersected with current files; new files selected by default
                saved = self._selected_files.get(key, set())
                if saved:
                    selected = set(f for f in files if f in saved)
                    # select also new files by default
                    for f in files:
                        if f not in saved:
                            selected.add(f)
                else:
                    selected = set(files)
                self._selected_files[key] = selected
                # render checkboxes
                chk_container = ctk.CTkFrame(sec, corner_radius=8)
                chk_container.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=6, pady=(4, 6))
                # arrange in two columns if many files
                for i in range(2):
                    chk_container.grid_columnconfigure(i, weight=1)
                for i, fname in enumerate(files):
                    col = i % 2
                    rowc = i // 2
                    var = ctk.BooleanVar(value=fname in selected)
                    def _make_cmd(k=key, name=fname, v=var):
                        return lambda: self._on_file_toggle(k, name, v.get())
                    cb = ctk.CTkCheckBox(chk_container, text=fname, variable=var, command=_make_cmd())
                    cb.grid(row=rowc, column=col, sticky="w", padx=6, pady=2)
            # raw_data controls: Send Minio / Save locally + path
            if key == "raw_data":
                has_checklist = bool(path and path.is_dir())
                next_row_local = 4 if has_checklist else 2
                # Send Minio checkbox
                send_var = ctk.BooleanVar(value=bool(self._raw_data_settings.get("send_minio", True)))
                def on_send_toggle():
                    self._raw_data_settings["send_minio"] = bool(send_var.get())
                    if callable(self.on_minio_toggle):
                        self.on_minio_toggle(send_var.get())
                    if callable(self.on_change):
                        self.on_change()
                send_cb = ctk.CTkCheckBox(sec, text="Send Minio", variable=send_var, command=on_send_toggle)
                send_cb.grid(row=next_row_local, column=0, sticky="w", padx=8, pady=(6, 4))
                # Save locally checkbox
                save_var = ctk.BooleanVar(value=bool(self._raw_data_settings.get("save_locally", False)))
                def on_save_toggle():
                    self._raw_data_settings["save_locally"] = bool(save_var.get())
                    entry.configure(state=("normal" if save_var.get() else "disabled"))
                    btn.configure(state=("normal" if save_var.get() else "disabled"))
                    if callable(self.on_change):
                        self.on_change()
                save_cb = ctk.CTkCheckBox(sec, text="Save locally", variable=save_var, command=on_save_toggle)
                save_cb.grid(row=next_row_local, column=1, sticky="w", padx=8, pady=(6, 4))
                # Path selector
                def choose_path():
                    folder = filedialog.askdirectory()
                    if folder:
                        entry.delete(0, "end")
                        entry.insert(0, folder)
                        self._raw_data_settings["local_path"] = folder
                        if callable(self.on_change):
                            self.on_change()
                ctk.CTkLabel(sec, text="Local path").grid(row=next_row_local + 1, column=0, sticky="w", padx=8, pady=4)
                entry = ctk.CTkEntry(sec, placeholder_text="Select a folder…")
                entry.grid(row=next_row_local + 1, column=1, sticky="ew", padx=(6, 8), pady=4)
                if self._raw_data_settings.get("local_path"):
                    entry.delete(0, "end")
                    entry.insert(0, self._raw_data_settings.get("local_path", ""))
                btn = ctk.CTkButton(sec, text="Browse…", width=90, command=choose_path)
                btn.grid(row=next_row_local + 2, column=1, sticky="e", padx=(6, 8), pady=(0, 6))
                entry.configure(state=("normal" if save_var.get() else "disabled"))
                btn.configure(state=("normal" if save_var.get() else "disabled"))
            # metrics DataFrame controls
            if key == "metrics" and path and path.is_file():
                col_names, data_rows = self._read_tabular(path, sheet)
                # defaults for selected columns: if none saved, select all except time col
                if not self._metrics_settings.get("selected_cols"):
                    self._metrics_settings["selected_cols"] = set(col_names)
                # Header checkbox
                header_var = ctk.BooleanVar(value=bool(self._metrics_settings.get("header", True)))
                def on_header_toggle():
                    self._metrics_settings["header"] = bool(header_var.get())
                    # reset selected cols to match new headers
                    cols2, _ = self._read_tabular(path, sheet)
                    self._metrics_settings["selected_cols"] = set(cols2)
                    # if time col no longer exists, reset
                    if self._metrics_settings.get("time_col") not in cols2:
                        self._metrics_settings["time_col"] = ""
                    self.render_details_sections()
                    if callable(self.on_change):
                        self.on_change()
                header_cb = ctk.CTkCheckBox(sec, text="Column header", variable=header_var, command=on_header_toggle)
                header_cb.grid(row=3, column=0, sticky="w", padx=8, pady=(6, 4))

                # Time column checkbox
                has_time_var = ctk.BooleanVar(value=bool(self._metrics_settings.get("has_time", False)))
                def on_has_time_toggle():
                    self._metrics_settings["has_time"] = bool(has_time_var.get())
                    if not has_time_var.get():
                        self._metrics_settings["time_col"] = ""
                    self.render_details_sections()
                    if callable(self.on_change):
                        self.on_change()
                time_cb = ctk.CTkCheckBox(sec, text="x-axis column", variable=has_time_var, command=on_has_time_toggle)
                time_cb.grid(row=3, column=1, sticky="w", padx=8, pady=(6, 4))

                next_row = 4
                current_cols = list(col_names)
                # Time column selector if enabled
                if has_time_var.get():
                    time_values = current_cols
                    time_menu = ctk.CTkOptionMenu(sec, values=time_values, dynamic_resizing=False,
                                                  command=lambda v: self._on_metrics_time_column_changed(v))
                    # set current
                    if self._metrics_settings.get("time_col") in time_values:
                        time_menu.set(self._metrics_settings.get("time_col"))
                    elif time_values:
                        time_menu.set(time_values[0])
                        self._metrics_settings["time_col"] = time_values[0]
                    ctk.CTkLabel(sec, text="x-axis column").grid(row=next_row, column=0, sticky="w", padx=8, pady=4)
                    time_menu.grid(row=next_row, column=1, sticky="ew", padx=(6, 8), pady=4)
                    next_row += 1

                # Columns checklist (exclude time column if set)
                cols_to_list = [c for c in current_cols if c != self._metrics_settings.get("time_col", "")]
                chk_container = ctk.CTkFrame(sec, corner_radius=8)
                chk_container.grid(row=next_row, column=0, columnspan=2, sticky="nsew", padx=6, pady=(4, 6))
                for i in range(2):
                    chk_container.grid_columnconfigure(i, weight=1)
                for i, cname in enumerate(cols_to_list):
                    col = i % 2
                    rowc = i // 2
                    var = ctk.BooleanVar(value=(cname in self._metrics_settings.get("selected_cols", set())))
                    def _make_cmd2(name=cname, v=var):
                        return lambda: self._on_metrics_column_toggle(name, v.get())
                    cb = ctk.CTkCheckBox(chk_container, text=cname, variable=var, command=_make_cmd2())
                    cb.grid(row=rowc, column=col, sticky="w", padx=6, pady=2)
            idx += 1

    def _on_file_toggle(self, key: str, filename: str, is_selected: bool):
        sel = self._selected_files.get(key, set())
        if is_selected:
            sel.add(filename)
        else:
            sel.remove(filename)
        self._selected_files[key] = sel
        if callable(self.on_change):
            self.on_change()

    # --- Metrics helpers ---
    def _read_tabular(self, path: Path, sheet: str) -> tuple[list[str], list[list[object]]]:
        cols: list[str] = []
        rows: list[list[object]] = []
        try:
            if path.suffix.lower() in (".xlsx", ".xlsm"):
                wb = load_workbook(filename=str(path), read_only=True, data_only=True)
                ws = None
                if sheet and sheet in wb.sheetnames:
                    ws = wb[sheet]
                else:
                    ws = wb[wb.sheetnames[0]]
                data_iter = ws.iter_rows(values_only=True)
                for i, row in enumerate(data_iter):
                    if i == 0 and bool(self._metrics_settings.get("header", True)):
                        cols = [str(c) if c is not None else f"col{idx}" for idx, c in enumerate(list(row))]
                    else:
                        rows.append(list(row))
                wb.close()
                # if no header, generate from max row length
                if not cols:
                    max_len = max((len(r) for r in rows), default=0)
                    cols = [str(i) for i in range(max_len)]
            elif path.suffix.lower() == ".csv":
                # Use selected separator for metrics preview
                sep = self._csv_separators.get("metrics", ",")
                with open(path, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f, delimiter=("\t" if sep == "\t" else sep))
                    for i, row in enumerate(reader):
                        if i == 0 and bool(self._metrics_settings.get("header", True)):
                            cols = [str(c) for c in row]
                        else:
                            rows.append(row)
                if not cols:
                    max_len = max((len(r) for r in rows), default=0)
                    cols = [str(i) for i in range(max_len)]
            else:
                # unsupported -> empty
                cols, rows = [], []
        except Exception as e:
            _log_error(e, "Error reading tabular data")
            self.status.configure(text=f"❌ Error reading metrics: {e}")
        return cols, rows

    def _on_sep_changed(self, key: str, display_value: str):
        sep = "\t" if display_value == "\\t" else display_value
        self._csv_separators[key] = sep
        # re-render metrics preview immediately to reflect new parsing
        if key == "metrics":
            self.render_details_sections()
        if callable(self.on_change):
            self.on_change()

    def _on_metrics_time_column_changed(self, col_name: str):
        self._metrics_settings["time_col"] = col_name or ""
        if col_name:
            # ensure selected cols include all except time col if none set
            if not self._metrics_settings.get("selected_cols"):
                self._metrics_settings["selected_cols"] = set()
        self.render_details_sections()
        if callable(self.on_change):
            self.on_change()

    def _on_metrics_column_toggle(self, col_name: str, is_selected: bool):
        sel = self._metrics_settings.get("selected_cols", set())
        if is_selected:
            sel.add(col_name)
        else:
            try:
                sel.remove(col_name)
            except KeyError:
                pass
        self._metrics_settings["selected_cols"] = sel
        if callable(self.on_change):
            self.on_change()

    def _flatten_dict(self, d: dict, parent_key: str = "", sep: str = "_") -> dict:
        """Flatten a nested dictionary with keys joined by separator."""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep).items())
            elif isinstance(v, list):
                # For lists, just keep them as-is
                items.append((new_key, v))
            else:
                items.append((new_key, v))
        return dict(items)

    def _read_config_preview(self, path: Path, sheet: str = "", max_lines: int = 12) -> str:
        """Read config file and return a formatted preview string."""
        try:
            suffix = path.suffix.lower()
            if suffix == ".json":
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                # Apply flatten if enabled
                if self._config_settings.get("flatten", False):
                    data = self._flatten_dict(data)
                # Format JSON with indentation
                preview = json.dumps(data, indent=2, ensure_ascii=False)
            elif suffix in (".yaml", ".yml"):
                if yaml is None:
                    preview = "(PyYAML not installed - run: pip install pyyaml)"
                else:
                    with open(path, "r", encoding="utf-8") as f:
                        data = yaml.safe_load(f)
                    # Apply flatten if enabled
                    if self._config_settings.get("flatten", False) and isinstance(data, dict):
                        data = self._flatten_dict(data)
                    # Format as YAML
                    preview = yaml.dump(data, default_flow_style=False, allow_unicode=True, sort_keys=False)
            elif suffix == ".csv":
                sep = self._csv_separators.get("config", ",")
                sep = "\t" if sep in ("\\t", "\t") else sep
                lines = []
                with open(path, "r", newline="", encoding="utf-8") as f:
                    reader = csv.reader(f, delimiter=sep)
                    for row in reader:
                        lines.append(" │ ".join(str(c) for c in row))
                preview = "\n".join(lines)
            elif suffix in (".xlsx", ".xlsm"):
                wb = load_workbook(filename=str(path), read_only=True, data_only=True)
                ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
                lines = []
                for row in ws.iter_rows(values_only=True):
                    row_str = " │ ".join(str(c) if c is not None else "" for c in row)
                    lines.append(row_str)
                wb.close()
                preview = "\n".join(lines)
            else:
                preview = "(Unsupported format)"
            
            # Truncate if too long
            lines = preview.split("\n")
            if len(lines) > max_lines:
                preview = "\n".join(lines[:max_lines]) + f"\n... ({len(lines) - max_lines} more lines)"
            
            # Also limit total characters
            if len(preview) > 1500:
                preview = preview[:1500] + "\n... (truncated)"
            
            return preview
        except Exception as e:
            return f"(Error reading config: {e})"

